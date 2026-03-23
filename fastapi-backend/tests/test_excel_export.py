"""Unit tests for Excel export service (app.ai.excel_export).

Tests cover:
- Workbook creation and formatting
- Header styling (bold, blue fill)
- Data row integrity
- Empty rows (headers only)
- Filename sanitisation
- Download URL pattern
- File lookup (get_export_path)
- Path traversal rejection
- Stale file cleanup (_cleanup_old_exports)
"""

from __future__ import annotations

import os
import shutil
import time
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.ai.excel_export import (
    EXPORT_DIR,
    EXPORT_TTL_SECONDS,
    ExportResult,
    _cleanup_old_exports,
    export_to_excel,
    get_export_path,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _user_dir(user_id) -> Path:
    return EXPORT_DIR / str(user_id)


def _cleanup_user(user_id) -> None:
    """Remove a user's export directory after a test."""
    d = _user_dir(user_id)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)


# ---------------------------------------------------------------------------
# Tests: export_to_excel
# ---------------------------------------------------------------------------


class TestExportToExcel:
    async def test_creates_valid_xlsx(self):
        uid = uuid4()
        try:
            result = await export_to_excel(
                columns=["name", "age"],
                rows=[{"name": "Alice", "age": 30}],
                title="People",
                user_id=uid,
            )

            assert isinstance(result, ExportResult)

            file_path = _user_dir(uid) / result.filename
            assert file_path.is_file()

            wb = load_workbook(str(file_path))
            ws = wb.active
            assert ws is not None
            wb.close()
        finally:
            _cleanup_user(uid)

    async def test_header_row_bold_blue(self):
        uid = uuid4()
        try:
            result = await export_to_excel(
                columns=["col_a", "col_b"],
                rows=[{"col_a": 1, "col_b": 2}],
                title="Styled",
                user_id=uid,
            )
            file_path = _user_dir(uid) / result.filename
            wb = load_workbook(str(file_path))
            ws = wb.active

            for col_idx in range(1, 3):
                cell = ws.cell(row=1, column=col_idx)
                assert cell.font.bold is True
                # Header fill colour is #4472C4
                assert cell.fill.start_color.rgb is not None
                assert "4472C4" in str(cell.fill.start_color.rgb)

            wb.close()
        finally:
            _cleanup_user(uid)

    async def test_data_rows_match_input(self):
        uid = uuid4()
        columns = ["id", "value"]
        rows = [
            {"id": 1, "value": "one"},
            {"id": 2, "value": "two"},
            {"id": 3, "value": "three"},
        ]
        try:
            result = await export_to_excel(
                columns=columns,
                rows=rows,
                title="Data",
                user_id=uid,
            )
            file_path = _user_dir(uid) / result.filename
            wb = load_workbook(str(file_path))
            ws = wb.active

            # Verify each data row (rows start at 2)
            for row_idx, row_data in enumerate(rows, start=2):
                assert ws.cell(row=row_idx, column=1).value == row_data["id"]
                assert ws.cell(row=row_idx, column=2).value == row_data["value"]

            wb.close()
        finally:
            _cleanup_user(uid)

    async def test_empty_rows_only_headers(self):
        uid = uuid4()
        try:
            result = await export_to_excel(
                columns=["x", "y"],
                rows=[],
                title="Empty",
                user_id=uid,
            )

            assert result.row_count == 0

            file_path = _user_dir(uid) / result.filename
            wb = load_workbook(str(file_path))
            ws = wb.active

            # Headers present
            assert ws.cell(row=1, column=1).value == "x"
            assert ws.cell(row=1, column=2).value == "y"

            # No data rows
            assert ws.cell(row=2, column=1).value is None

            wb.close()
        finally:
            _cleanup_user(uid)

    async def test_filename_contains_sanitised_title(self):
        uid = uuid4()
        try:
            result = await export_to_excel(
                columns=["a"],
                rows=[{"a": 1}],
                title="My Report 2026!",
                user_id=uid,
            )
            # Special chars stripped, spaces become underscores
            assert "My_Report_2026" in result.filename
            assert result.filename.endswith(".xlsx")
        finally:
            _cleanup_user(uid)

    async def test_download_url_pattern(self):
        uid = uuid4()
        try:
            result = await export_to_excel(
                columns=["a"],
                rows=[{"a": 1}],
                title="Test",
                user_id=uid,
            )
            assert result.download_url.startswith("/api/ai/export/")
            assert result.filename in result.download_url
        finally:
            _cleanup_user(uid)

    async def test_row_count_matches(self):
        uid = uuid4()
        rows = [{"a": i} for i in range(5)]
        try:
            result = await export_to_excel(
                columns=["a"],
                rows=rows,
                title="Count",
                user_id=uid,
            )
            assert result.row_count == 5
        finally:
            _cleanup_user(uid)


# ---------------------------------------------------------------------------
# Tests: get_export_path
# ---------------------------------------------------------------------------


class TestGetExportPath:
    async def test_returns_path_for_existing_file(self):
        uid = uuid4()
        try:
            result = await export_to_excel(
                columns=["a"],
                rows=[{"a": 1}],
                title="Lookup",
                user_id=uid,
            )
            path = get_export_path(result.filename, uid)
            assert path is not None
            assert path.is_file()
        finally:
            _cleanup_user(uid)

    def test_returns_none_for_nonexistent_file(self):
        uid = uuid4()
        path = get_export_path("does_not_exist.xlsx", uid)
        assert path is None

    @pytest.mark.parametrize(
        "bad_name",
        [
            "../../etc/passwd",
            "../secret.xlsx",
            "..\\windows\\system32",
            "sub/dir/file.xlsx",
            "sub\\dir\\file.xlsx",
        ],
    )
    def test_rejects_path_traversal(self, bad_name):
        uid = uuid4()
        path = get_export_path(bad_name, uid)
        assert path is None


# ---------------------------------------------------------------------------
# Tests: _cleanup_old_exports
# ---------------------------------------------------------------------------


class TestCleanupOldExports:
    async def test_removes_old_files(self):
        uid = uuid4()
        try:
            # Create a file
            result = await export_to_excel(
                columns=["a"],
                rows=[{"a": 1}],
                title="Old",
                user_id=uid,
            )
            file_path = _user_dir(uid) / result.filename
            assert file_path.is_file()

            # Set mtime to 2 hours ago (past TTL)
            old_time = time.time() - EXPORT_TTL_SECONDS - 3600
            os.utime(str(file_path), (old_time, old_time))

            removed = _cleanup_old_exports()
            assert removed >= 1
            assert not file_path.is_file()
        finally:
            _cleanup_user(uid)

    def test_no_error_when_dir_missing(self):
        """Cleanup returns 0 when EXPORT_DIR doesn't exist (no crash)."""
        with patch("app.ai.excel_export.EXPORT_DIR", Path("/nonexistent_dir_12345")):
            assert _cleanup_old_exports() == 0

    async def test_keeps_recent_files(self):
        uid = uuid4()
        try:
            result = await export_to_excel(
                columns=["a"],
                rows=[{"a": 1}],
                title="Recent",
                user_id=uid,
            )
            file_path = _user_dir(uid) / result.filename

            _cleanup_old_exports()
            # File was just created, should still exist
            assert file_path.is_file()
        finally:
            _cleanup_user(uid)
