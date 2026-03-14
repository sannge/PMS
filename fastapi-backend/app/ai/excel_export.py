"""Excel export service for AI query results.

Creates formatted .xlsx files from query results using openpyxl.
Files stored in temp directory scoped by user_id with 1-hour TTL cleanup.
"""

from __future__ import annotations

import asyncio
import logging
import tempfile
import time
from dataclasses import dataclass
from functools import partial
from pathlib import Path
from uuid import UUID, uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config_service import get_agent_config

logger = logging.getLogger(__name__)

_cfg = get_agent_config()

EXPORT_DIR = Path(tempfile.gettempdir()) / "blair_exports"
EXPORT_TTL_SECONDS = _cfg.get_int("export.excel_ttl_seconds", 3600)

# Header styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Minimum column width (characters)
MIN_COL_WIDTH = 10
# Maximum column width
MAX_COL_WIDTH = 50


@dataclass
class ExportResult:
    """Result of Excel export."""

    filename: str
    download_url: str
    row_count: int


def _cleanup_old_exports() -> int:
    """Remove export files older than EXPORT_TTL_SECONDS.

    Returns:
        Number of files removed.
    """
    if not EXPORT_DIR.exists():
        return 0

    removed = 0
    now = time.time()

    for user_dir in EXPORT_DIR.iterdir():
        if not user_dir.is_dir():
            continue
        for file_path in user_dir.iterdir():
            if not file_path.is_file():
                continue
            try:
                age = now - file_path.stat().st_mtime
                if age > EXPORT_TTL_SECONDS:
                    file_path.unlink()
                    removed += 1
            except OSError:
                logger.debug("Failed to remove expired export: %s", file_path)

        # Remove empty user directories
        try:
            if user_dir.is_dir() and not any(user_dir.iterdir()):
                user_dir.rmdir()
        except OSError:
            pass

    if removed:
        logger.info("Cleaned up %d expired export files", removed)
    return removed


async def export_to_excel(
    columns: list[str],
    rows: list[dict],
    title: str,
    user_id: UUID,
) -> ExportResult:
    """Create an Excel workbook from query results.

    Creates a formatted .xlsx file with:
    - Bold header row with blue background
    - Auto-sized columns based on content width
    - Data rows with proper alignment

    Files are stored in a user-scoped temp directory and cleaned up
    after EXPORT_TTL_SECONDS.

    Args:
        columns: Column names for the header row.
        rows: List of row dicts (keys matching columns).
        title: Title for the export (used in filename).
        user_id: User UUID for file scoping.

    Returns:
        ExportResult with filename, download URL, and row count.
    """
    # Clean up old exports first (offload blocking FS traversal to thread)
    await asyncio.to_thread(_cleanup_old_exports)

    # Create user-scoped directory
    user_dir = EXPORT_DIR / str(user_id)
    user_dir.mkdir(parents=True, exist_ok=True)

    # Generate unique filename
    safe_title = "".join(c if c.isalnum() or c in "-_ " else "" for c in title)
    safe_title = safe_title.strip().replace(" ", "_")[:50] or "export"
    filename = f"{safe_title}_{uuid4().hex[:8]}.xlsx"
    file_path = user_dir / filename

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = safe_title[:31]  # Excel sheet name max 31 chars

    # Write header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns based on content
    for col_idx, col_name in enumerate(columns, start=1):
        # Start with header width
        max_width = len(str(col_name))

        # Check data widths (sample first 100 rows)
        for row_idx in range(2, min(len(rows) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_width = max(max_width, len(str(cell_value)))

        # Apply width with bounds
        adjusted_width = min(max(max_width + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save workbook (offload blocking I/O to thread pool)
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, partial(wb.save, str(file_path)))

    download_url = f"/api/ai/export/{filename}"

    logger.info(
        "Excel export created: %s (%d rows) for user %s",
        filename,
        len(rows),
        user_id,
    )

    return ExportResult(
        filename=filename,
        download_url=download_url,
        row_count=len(rows),
    )


def get_export_path(filename: str, user_id: UUID) -> Path | None:
    """Get the file path for an export if it exists and belongs to the user.

    Validates that the filename is safe and the file exists within the
    user's export directory. Prevents path traversal attacks.

    Args:
        filename: The export filename to look up.
        user_id: The user UUID to scope the lookup.

    Returns:
        Path to the file if it exists and belongs to the user, None otherwise.
    """
    # Sanitize filename to prevent path traversal
    safe_name = Path(filename).name
    if safe_name != filename or ".." in filename or "/" in filename or "\\" in filename:
        logger.warning("Rejected suspicious export filename: %s", filename)
        return None

    file_path = EXPORT_DIR / str(user_id) / safe_name

    # Verify the resolved path is within the expected directory
    try:
        file_path.resolve().relative_to(EXPORT_DIR.resolve())
    except ValueError:
        logger.warning("Path traversal attempt detected: %s", filename)
        return None

    if file_path.is_file():
        return file_path

    return None
