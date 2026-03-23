"""Team Activity aggregation service.

Provides data for the Team Activity page: overview KPIs, member breakdowns,
project breakdowns, activity feed, and Excel export.
"""

from __future__ import annotations

import logging
import os
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path
from uuid import UUID

from sqlalchemy import case, func, literal, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.application import Application
from ..models.application_member import ApplicationMember
from ..models.comment import Comment
from ..models.document import Document
from ..models.project import Project
from ..models.project_member import ProjectMember
from ..models.task import Task
from ..models.task_status import TaskStatus
from ..models.user import User
from ..schemas.team_activity import (
    KPICards,
    MemberBreakdown,
    MemberDetailResponse,
    MemberDocDetail,
    MembersSummaryResponse,
    MemberSummary,
    MemberTaskDetail,
    OverviewResponse,
    ProjectBreakdown,
    ProjectDetailResponse,
    ProjectMemberBreakdown,
    ProjectsSummaryResponse,
    ProjectSummary,
    ProjectTaskRow,
    WeeklyCompletion,
)
from ..utils.timezone import utc_now

logger = logging.getLogger(__name__)

# Redis cache TTL for overview data (seconds)
CACHE_TTL = 60

# Statement timeout for all queries (milliseconds)
_STATEMENT_TIMEOUT_MS = "3000"


async def _set_statement_timeout(db: AsyncSession) -> None:
    """Set a per-transaction statement timeout to prevent runaway queries."""
    await db.execute(text(f"SET LOCAL statement_timeout = '{_STATEMENT_TIMEOUT_MS}'"))


# ============================================================================
# Scope helpers
# ============================================================================


async def get_owned_app_ids(db: AsyncSession, user_id: UUID) -> list[UUID]:
    """Get all application IDs where user is owner."""
    result = await db.execute(
        select(Application.id).where(Application.owner_id == user_id)
    )
    return [row[0] for row in result.all()]


async def get_all_project_ids_for_apps(
    db: AsyncSession,
    app_ids: list[UUID],
) -> list[UUID]:
    """Get ALL project IDs for owned apps (no date filter — for RBAC checks)."""
    result = await db.execute(
        select(Project.id).where(Project.application_id.in_(app_ids))
    )
    return [row[0] for row in result.all()]


async def get_project_ids_for_apps(
    db: AsyncSession,
    app_ids: list[UUID],
    date_from: date,
    date_to: date,
) -> list[UUID]:
    """Get project IDs including archived projects with activity in range.

    Includes non-archived projects AND archived projects that had task
    completions within the date range (so archived work still shows in reports).
    """
    date_from_utc = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)

    result = await db.execute(
        select(Project.id).where(
            Project.application_id.in_(app_ids),
            # Include non-archived OR archived-but-had-completions-in-range
            (Project.archived_at.is_(None))
            | (
                (Project.archived_at.is_not(None))
                & (Project.archived_at >= date_from_utc)
            ),
        )
    )
    return [row[0] for row in result.all()]


# ============================================================================
# Overview
# ============================================================================


async def get_overview(
    db: AsyncSession,
    project_ids: list[UUID],
    date_from: date,
    date_to: date,
) -> OverviewResponse:
    """Build the overview tab data: KPIs, completion trend, breakdowns."""
    await _set_statement_timeout(db)

    date_from_utc = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)
    date_to_utc = datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc)
    now = utc_now()

    if not project_ids:
        return OverviewResponse(
            kpi=KPICards(
                completed=0, in_progress=0, in_review=0, overdue=0, total_story_points=0
            ),
            completion_trend=[],
            by_project=[],
            by_member=[],
        )

    # -- KPI cards: single query with conditional aggregation --
    kpi_result = await db.execute(
        select(
            func.count(Task.id).filter(
                TaskStatus.category == "Done",
                Task.completed_at >= date_from_utc,
                Task.completed_at <= date_to_utc,
            ).label("completed"),
            func.count(Task.id).filter(
                TaskStatus.name == "In Progress",
                Task.archived_at.is_(None),
            ).label("in_progress"),
            func.count(Task.id).filter(
                TaskStatus.name == "In Review",
                Task.archived_at.is_(None),
            ).label("in_review"),
            func.count(Task.id).filter(
                Task.due_date < now.date(),
                TaskStatus.category != "Done",
                Task.archived_at.is_(None),
            ).label("overdue"),
            func.coalesce(
                func.sum(Task.story_points).filter(
                    TaskStatus.category == "Done",
                    Task.completed_at >= date_from_utc,
                    Task.completed_at <= date_to_utc,
                ),
                0,
            ).label("total_story_points"),
        )
        .join(TaskStatus, Task.task_status_id == TaskStatus.id)
        .where(Task.project_id.in_(project_ids))
    )
    kpi_row = kpi_result.one()
    kpi = KPICards(
        completed=kpi_row.completed,
        in_progress=kpi_row.in_progress,
        in_review=kpi_row.in_review,
        overdue=kpi_row.overdue,
        total_story_points=kpi_row.total_story_points,
    )

    # -- Weekly completion trend --
    # Truncate completed_at to ISO week (Monday) and count
    week_expr = func.date_trunc("week", Task.completed_at)
    trend_result = await db.execute(
        select(
            week_expr.label("week_start"),
            func.count(Task.id).label("cnt"),
        )
        .join(TaskStatus, Task.task_status_id == TaskStatus.id)
        .where(
            Task.project_id.in_(project_ids),
            TaskStatus.category == "Done",
            Task.completed_at >= date_from_utc,
            Task.completed_at <= date_to_utc,
        )
        .group_by(week_expr)
        .order_by(week_expr)
    )
    completion_trend = [
        WeeklyCompletion(
            week=row.week_start.date().isoformat() if isinstance(row.week_start, datetime) else str(row.week_start),
            count=row.cnt,
        )
        for row in trend_result.all()
    ]

    # -- Per-project breakdown --
    proj_result = await db.execute(
        select(
            Project.id.label("project_id"),
            Project.name.label("project_name"),
            Project.key.label("project_key"),
            func.count(Task.id).filter(
                TaskStatus.category == "Done",
                Task.completed_at >= date_from_utc,
                Task.completed_at <= date_to_utc,
            ).label("completed"),
            func.count(Task.id).filter(
                TaskStatus.name == "In Progress",
                Task.archived_at.is_(None),
            ).label("in_progress"),
            func.count(Task.id).filter(
                TaskStatus.name == "In Review",
                Task.archived_at.is_(None),
            ).label("in_review"),
            func.count(Task.id).filter(
                Task.due_date < now.date(),
                TaskStatus.category != "Done",
                Task.archived_at.is_(None),
            ).label("overdue"),
            func.count(Task.id).filter(
                TaskStatus.category == "Todo",
                Task.archived_at.is_(None),
            ).label("todo"),
        )
        .join(Task, Task.project_id == Project.id)
        .join(TaskStatus, Task.task_status_id == TaskStatus.id)
        .where(Project.id.in_(project_ids))
        .group_by(Project.id, Project.name, Project.key)
        .order_by(func.count(Task.id).filter(
            TaskStatus.category == "Done",
            Task.completed_at >= date_from_utc,
            Task.completed_at <= date_to_utc,
        ).desc())
    )
    by_project = [
        ProjectBreakdown(
            project_id=str(r.project_id),
            project_name=r.project_name,
            project_key=r.project_key,
            completed=r.completed,
            in_progress=r.in_progress,
            in_review=r.in_review,
            overdue=r.overdue,
            todo=r.todo,
        )
        for r in proj_result.all()
    ]

    # -- Per-member breakdown (assignee) --
    member_result = await db.execute(
        select(
            User.id.label("user_id"),
            User.display_name,
            User.email,
            func.count(Task.id).filter(
                TaskStatus.category == "Done",
                Task.completed_at >= date_from_utc,
                Task.completed_at <= date_to_utc,
            ).label("completed"),
            func.count(Task.id).filter(
                TaskStatus.name == "In Progress",
                Task.archived_at.is_(None),
            ).label("in_progress"),
        )
        .join(Task, Task.assignee_id == User.id)
        .join(TaskStatus, Task.task_status_id == TaskStatus.id)
        .where(Task.project_id.in_(project_ids))
        .group_by(User.id, User.display_name, User.email)
        .order_by(func.count(Task.id).filter(
            TaskStatus.category == "Done",
            Task.completed_at >= date_from_utc,
            Task.completed_at <= date_to_utc,
        ).desc())
    )
    by_member = [
        MemberBreakdown(
            user_id=str(r.user_id),
            display_name=r.display_name or r.email,
            email=r.email,
            completed=r.completed,
            in_progress=r.in_progress,
        )
        for r in member_result.all()
    ]

    return OverviewResponse(
        kpi=kpi,
        completion_trend=completion_trend,
        by_project=by_project,
        by_member=by_member,
    )


async def get_overview_cached(
    db: AsyncSession,
    app_id_key: str,
    project_ids: list[UUID],
    date_from: date,
    date_to: date,
) -> OverviewResponse:
    """Get overview data with Redis caching (60s TTL)."""
    from .redis_service import redis_service

    cache_key = f"ta:overview:{app_id_key}:{date_from}:{date_to}"
    if redis_service.is_connected:
        try:
            cached = await redis_service.get(cache_key)
            if cached:
                return OverviewResponse.model_validate_json(cached)
        except Exception:
            pass  # Redis down -- fall through to DB

    result = await get_overview(db, project_ids, date_from, date_to)

    if redis_service.is_connected:
        try:
            await redis_service.set(cache_key, result.model_dump_json(), ttl=CACHE_TTL)
        except Exception:
            pass  # Redis down -- no caching

    return result


# ============================================================================
# Members
# ============================================================================


async def get_members_summary(
    db: AsyncSession,
    app_ids: list[UUID],
    project_ids: list[UUID],
    date_from: date,
    date_to: date,
) -> MembersSummaryResponse:
    """Aggregate counts per member across all owned apps."""
    await _set_statement_timeout(db)

    date_from_utc = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)
    date_to_utc = datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc)

    if not project_ids:
        return MembersSummaryResponse(members=[])

    # Get all members: owners + application members
    owner_ids = select(Application.owner_id.label("user_id")).where(
        Application.id.in_(app_ids),
        Application.owner_id.is_not(None),
    )
    app_member_ids = select(ApplicationMember.user_id.label("user_id")).where(
        ApplicationMember.application_id.in_(app_ids)
    )
    all_member_ids = owner_ids.union(app_member_ids).subquery()

    result = await db.execute(
        select(
            User.id.label("user_id"),
            User.display_name,
            User.email,
            User.avatar_url,
            # Role: check if owner of any app, else get app member role
            case(
                (
                    User.id.in_(
                        select(Application.owner_id).where(Application.id.in_(app_ids))
                    ),
                    literal("owner"),
                ),
                else_=func.coalesce(
                    select(ApplicationMember.role)
                    .where(
                        ApplicationMember.user_id == User.id,
                        ApplicationMember.application_id.in_(app_ids),
                    )
                    .correlate(User)
                    .limit(1)
                    .scalar_subquery(),
                    literal("member"),
                ),
            ).label("role"),
            # Task counts via subqueries
            select(func.count(Task.id))
            .join(TaskStatus, Task.task_status_id == TaskStatus.id)
            .where(
                Task.assignee_id == User.id,
                Task.project_id.in_(project_ids),
                TaskStatus.category == "Done",
                Task.completed_at >= date_from_utc,
                Task.completed_at <= date_to_utc,
            )
            .correlate(User)
            .scalar_subquery()
            .label("done_count"),
            select(func.count(Task.id))
            .join(TaskStatus, Task.task_status_id == TaskStatus.id)
            .where(
                Task.assignee_id == User.id,
                Task.project_id.in_(project_ids),
                TaskStatus.name == "In Progress",
                Task.archived_at.is_(None),
            )
            .correlate(User)
            .scalar_subquery()
            .label("in_progress_count"),
            select(func.count(Task.id))
            .join(TaskStatus, Task.task_status_id == TaskStatus.id)
            .where(
                Task.assignee_id == User.id,
                Task.project_id.in_(project_ids),
                TaskStatus.name == "In Review",
                Task.archived_at.is_(None),
            )
            .correlate(User)
            .scalar_subquery()
            .label("in_review_count"),
            func.coalesce(
                select(func.sum(Task.story_points))
                .join(TaskStatus, Task.task_status_id == TaskStatus.id)
                .where(
                    Task.assignee_id == User.id,
                    Task.project_id.in_(project_ids),
                    TaskStatus.category == "Done",
                    Task.completed_at >= date_from_utc,
                    Task.completed_at <= date_to_utc,
                )
                .correlate(User)
                .scalar_subquery(),
                0,
            ).label("story_points_sum"),
            # Document count: docs created by this user in owned apps
            select(func.count(Document.id))
            .where(
                Document.created_by == User.id,
                Document.deleted_at.is_(None),
                (
                    Document.application_id.in_(app_ids)
                    | Document.project_id.in_(project_ids)
                ),
                Document.created_at >= date_from_utc,
                Document.created_at <= date_to_utc,
            )
            .correlate(User)
            .scalar_subquery()
            .label("docs_count"),
            # Comment count
            select(func.count(Comment.id))
            .join(Task, Comment.task_id == Task.id)
            .where(
                Comment.author_id == User.id,
                Task.project_id.in_(project_ids),
                Comment.is_deleted.is_(False),
                Comment.created_at >= date_from_utc,
                Comment.created_at <= date_to_utc,
            )
            .correlate(User)
            .scalar_subquery()
            .label("comments_count"),
        )
        .where(User.id.in_(select(all_member_ids.c.user_id)))
        .order_by(User.display_name)
    )

    members = [
        MemberSummary(
            user_id=str(r.user_id),
            display_name=r.display_name or r.email,
            email=r.email,
            avatar_url=r.avatar_url,
            role=r.role,
            done_count=r.done_count or 0,
            in_progress_count=r.in_progress_count or 0,
            in_review_count=r.in_review_count or 0,
            story_points_sum=r.story_points_sum or 0,
            docs_count=r.docs_count or 0,
            comments_count=r.comments_count or 0,
        )
        for r in result.all()
    ]

    return MembersSummaryResponse(members=members)


async def get_member_detail(
    db: AsyncSession,
    member_user_id: UUID,
    app_ids: list[UUID],
    project_ids: list[UUID],
    date_from: date,
    date_to: date,
) -> MemberDetailResponse:
    """Get tasks + docs for one member (lazy-loaded on expand)."""
    await _set_statement_timeout(db)

    date_from_utc = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)
    date_to_utc = datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc)

    # Tasks assigned to this user in scoped projects
    task_result = await db.execute(
        select(
            Task.id.label("task_id"),
            Task.task_key,
            Task.title,
            Project.name.label("project_name"),
            Project.key.label("project_key"),
            TaskStatus.name.label("status_name"),
            TaskStatus.category.label("status_category"),
            Task.priority,
            Task.story_points,
            Task.completed_at,
            Task.created_at,
        )
        .join(TaskStatus, Task.task_status_id == TaskStatus.id)
        .join(Project, Task.project_id == Project.id)
        .where(
            Task.assignee_id == member_user_id,
            Task.project_id.in_(project_ids),
            # Include tasks active now OR completed in range
            (Task.archived_at.is_(None))
            | (
                (Task.completed_at >= date_from_utc)
                & (Task.completed_at <= date_to_utc)
            ),
        )
        .order_by(Task.updated_at.desc())
        .limit(200)
    )
    tasks = [
        MemberTaskDetail(
            task_id=str(r.task_id),
            task_key=r.task_key,
            title=r.title,
            project_name=r.project_name,
            project_key=r.project_key,
            status_name=r.status_name,
            status_category=r.status_category,
            priority=r.priority,
            story_points=r.story_points,
            completed_at=r.completed_at.isoformat() if r.completed_at else None,
            created_at=r.created_at.isoformat(),
        )
        for r in task_result.all()
    ]

    # Documents created by this user in scoped apps/projects
    doc_result = await db.execute(
        select(
            Document.id.label("document_id"),
            Document.title,
            Document.application_id,
            Document.project_id,
            Document.user_id.label("doc_user_id"),
            case(
                (Document.application_id.is_not(None), literal("application")),
                (Document.project_id.is_not(None), literal("project")),
                else_=literal("personal"),
            ).label("scope"),
            func.coalesce(
                select(Application.name)
                .where(Application.id == Document.application_id)
                .correlate(Document)
                .scalar_subquery(),
                select(Project.name)
                .where(Project.id == Document.project_id)
                .correlate(Document)
                .scalar_subquery(),
                literal("Personal"),
            ).label("scope_name"),
            Document.created_at,
            Document.updated_at,
        )
        .where(
            Document.created_by == member_user_id,
            Document.deleted_at.is_(None),
            Document.created_at >= date_from_utc,
            Document.created_at <= date_to_utc,
            # Scope to owned apps/projects only — prevent cross-app data leak
            (
                Document.application_id.in_(app_ids)
                | Document.project_id.in_(project_ids)
            ),
        )
        .order_by(Document.updated_at.desc())
        .limit(100)
    )
    documents = [
        MemberDocDetail(
            document_id=str(r.document_id),
            title=r.title,
            scope=r.scope,
            scope_name=r.scope_name or "Unknown",
            created_at=r.created_at.isoformat(),
            updated_at=r.updated_at.isoformat(),
        )
        for r in doc_result.all()
    ]

    # Comment count
    comment_result = await db.execute(
        select(func.count(Comment.id))
        .join(Task, Comment.task_id == Task.id)
        .where(
            Comment.author_id == member_user_id,
            Task.project_id.in_(project_ids),
            Comment.is_deleted.is_(False),
            Comment.created_at >= date_from_utc,
            Comment.created_at <= date_to_utc,
        )
    )
    comments_count = comment_result.scalar() or 0

    return MemberDetailResponse(
        user_id=str(member_user_id),
        tasks=tasks,
        documents=documents,
        comments_count=comments_count,
    )


# ============================================================================
# Projects
# ============================================================================


async def get_projects_summary(
    db: AsyncSession,
    app_ids: list[UUID],
    date_from: date,
    date_to: date,
) -> ProjectsSummaryResponse:
    """Per-project status counts including archived."""
    await _set_statement_timeout(db)

    date_from_utc = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)

    if not app_ids:
        return ProjectsSummaryResponse(projects=[])

    result = await db.execute(
        select(
            Project.id.label("project_id"),
            Project.name.label("project_name"),
            Project.key.label("project_key"),
            Application.name.label("application_name"),
            Project.due_date,
            Project.archived_at,
            # Task counts using conditional aggregation
            func.count(Task.id).label("total"),
            func.count(Task.id).filter(
                TaskStatus.category == "Done",
            ).label("done"),
            func.count(Task.id).filter(
                TaskStatus.name == "In Progress",
                Task.archived_at.is_(None),
            ).label("in_progress"),
            func.count(Task.id).filter(
                TaskStatus.name == "In Review",
                Task.archived_at.is_(None),
            ).label("in_review"),
            func.count(Task.id).filter(
                TaskStatus.category == "Issue",
                Task.archived_at.is_(None),
            ).label("issue"),
            func.count(Task.id).filter(
                TaskStatus.category == "Todo",
                Task.archived_at.is_(None),
            ).label("todo"),
            func.count(Task.id).filter(
                Task.archived_at.is_not(None),
            ).label("archived_count"),
            func.count(Task.id).filter(
                Task.assignee_id.is_(None),
                Task.archived_at.is_(None),
            ).label("unassigned"),
        )
        .join(Application, Project.application_id == Application.id)
        .outerjoin(Task, Task.project_id == Project.id)
        .outerjoin(TaskStatus, Task.task_status_id == TaskStatus.id)
        .where(
            Project.application_id.in_(app_ids),
            # Include non-archived OR recently archived
            (Project.archived_at.is_(None))
            | (Project.archived_at >= date_from_utc),
        )
        .group_by(
            Project.id,
            Project.name,
            Project.key,
            Application.name,
            Project.due_date,
            Project.archived_at,
        )
        .order_by(Project.updated_at.desc())
    )

    projects: list[ProjectSummary] = []
    project_ids_for_members: list[UUID] = []
    for r in result.all():
        project_ids_for_members.append(r.project_id)
        total = r.total
        done = r.done
        pct = round((done / total) * 100, 1) if total > 0 else 0.0
        projects.append(
            ProjectSummary(
                project_id=str(r.project_id),
                project_name=r.project_name,
                project_key=r.project_key,
                application_name=r.application_name,
                due_date=r.due_date.isoformat() if r.due_date else None,
                total=total,
                done=done,
                in_progress=r.in_progress,
                in_review=r.in_review,
                issue=r.issue,
                todo=r.todo,
                archived=r.archived_count,
                unassigned=r.unassigned,
                is_archived=r.archived_at is not None,
                archived_at=r.archived_at.isoformat() if r.archived_at else None,
                members=[],  # filled below
                progress_pct=pct,
            )
        )

    # Fetch project members display names
    if project_ids_for_members:
        pm_result = await db.execute(
            select(
                ProjectMember.project_id,
                User.display_name,
                User.email,
            )
            .join(User, ProjectMember.user_id == User.id)
            .where(ProjectMember.project_id.in_(project_ids_for_members))
            .order_by(User.display_name)
        )
        members_by_project: dict[UUID, list[str]] = {}
        for pm_row in pm_result.all():
            members_by_project.setdefault(pm_row.project_id, []).append(
                pm_row.display_name or pm_row.email
            )
        for proj in projects:
            proj.members = members_by_project.get(UUID(proj.project_id), [])

    return ProjectsSummaryResponse(projects=projects)


async def get_project_detail(
    db: AsyncSession,
    project_id: UUID,
    date_from: date,
    date_to: date,
) -> ProjectDetailResponse:
    """Member breakdown + task list for one project (lazy-loaded)."""
    await _set_statement_timeout(db)

    date_from_utc = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)
    date_to_utc = datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc)

    # Task visibility filter: non-archived tasks (current state) OR
    # archived tasks where archived_at falls in range OR
    # tasks completed within the date range.
    task_date_filter = (
        (Task.archived_at.is_(None))
        | (
            (Task.archived_at >= date_from_utc)
            & (Task.archived_at <= date_to_utc)
        )
        | (
            (Task.completed_at >= date_from_utc)
            & (Task.completed_at <= date_to_utc)
        )
    )

    # Member breakdown by assignee
    mb_result = await db.execute(
        select(
            User.id.label("user_id"),
            User.display_name,
            func.count(Task.id).filter(
                TaskStatus.category == "Done",
                Task.completed_at >= date_from_utc,
                Task.completed_at <= date_to_utc,
            ).label("done"),
            func.count(Task.id).filter(TaskStatus.name == "In Progress").label("in_progress"),
            func.count(Task.id).filter(TaskStatus.name == "In Review").label("in_review"),
            func.count(Task.id).filter(TaskStatus.category == "Issue").label("issue"),
            func.count(Task.id).filter(TaskStatus.category == "Todo").label("todo"),
            func.coalesce(
                func.sum(Task.story_points).filter(
                    TaskStatus.category == "Done",
                    Task.completed_at >= date_from_utc,
                    Task.completed_at <= date_to_utc,
                ), 0
            ).label("story_points"),
        )
        .join(Task, Task.assignee_id == User.id)
        .join(TaskStatus, Task.task_status_id == TaskStatus.id)
        .where(Task.project_id == project_id, task_date_filter)
        .group_by(User.id, User.display_name)
        .order_by(User.display_name)
    )
    member_breakdown = [
        ProjectMemberBreakdown(
            user_id=str(r.user_id),
            display_name=r.display_name or "",
            done=r.done,
            in_progress=r.in_progress,
            in_review=r.in_review,
            issue=r.issue,
            todo=r.todo,
            story_points=r.story_points,
        )
        for r in mb_result.all()
    ]

    # Task list
    task_result = await db.execute(
        select(
            Task.id.label("task_id"),
            Task.task_key,
            Task.title,
            TaskStatus.name.label("status_name"),
            TaskStatus.category.label("status_category"),
            Task.priority,
            User.display_name.label("assignee_name"),
            Task.completed_at,
            Task.archived_at,
        )
        .join(TaskStatus, Task.task_status_id == TaskStatus.id)
        .outerjoin(User, Task.assignee_id == User.id)
        .where(Task.project_id == project_id, task_date_filter)
        .order_by(Task.updated_at.desc())
        .limit(500)
    )
    tasks = [
        ProjectTaskRow(
            task_id=str(r.task_id),
            task_key=r.task_key,
            title=r.title,
            status_name=r.status_name,
            status_category=r.status_category,
            priority=r.priority,
            assignee_name=r.assignee_name,
            completed_at=r.completed_at.isoformat() if r.completed_at else None,
            is_archived=r.archived_at is not None,
        )
        for r in task_result.all()
    ]

    return ProjectDetailResponse(
        project_id=str(project_id),
        member_breakdown=member_breakdown,
        tasks=tasks,
    )


# ============================================================================
# Export
# ============================================================================


def _build_workbook(
    overview_data: dict,
    members_data: dict,
    projects_data: dict,
    tab: str,
    date_from: str,
    date_to: str,
) -> str:
    """Synchronous workbook generation (runs in thread pool).

    Args:
        overview_data: Serialized OverviewResponse dict.
        members_data: Serialized MembersSummaryResponse dict.
        projects_data: Serialized ProjectsSummaryResponse dict.
        tab: Which tab(s) to export: "overview", "members", "projects", or "all".
        date_from: ISO date string for header.
        date_to: ISO date string for header.

    Returns:
        Path to the generated .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    def _write_headers(ws, headers: list[str]) -> None:  # type: ignore[no-untyped-def]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    sheets_created = False

    if tab in ("overview", "all"):
        ws = wb.active
        if ws is not None:
            ws.title = "Overview KPIs"
        else:
            ws = wb.create_sheet("Overview KPIs")
        sheets_created = True

        kpi = overview_data.get("kpi", {})
        _write_headers(ws, ["Metric", "Value"])
        metrics = [
            ("Completed", kpi.get("completed", 0)),
            ("In Progress", kpi.get("in_progress", 0)),
            ("In Review", kpi.get("in_review", 0)),
            ("Overdue", kpi.get("overdue", 0)),
            ("Total Story Points", kpi.get("total_story_points", 0)),
            ("Date Range", f"{date_from} to {date_to}"),
        ]
        for row_idx, (metric, value) in enumerate(metrics, 2):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        # Project breakdown sheet
        ws_proj = wb.create_sheet("By Project")
        _write_headers(ws_proj, [
            "Project", "Completed", "In Progress", "In Review", "Overdue", "Todo",
        ])
        for row_idx, proj in enumerate(overview_data.get("by_project", []), 2):
            ws_proj.cell(row=row_idx, column=1, value=proj.get("project_name", ""))
            ws_proj.cell(row=row_idx, column=2, value=proj.get("completed", 0))
            ws_proj.cell(row=row_idx, column=3, value=proj.get("in_progress", 0))
            ws_proj.cell(row=row_idx, column=4, value=proj.get("in_review", 0))
            ws_proj.cell(row=row_idx, column=5, value=proj.get("overdue", 0))
            ws_proj.cell(row=row_idx, column=6, value=proj.get("todo", 0))

    if tab in ("members", "all"):
        ws_m = wb.create_sheet("Members") if sheets_created else wb.active
        if ws_m is not None and not sheets_created:
            ws_m.title = "Members"
        elif ws_m is None:
            ws_m = wb.create_sheet("Members")
        sheets_created = True

        _write_headers(ws_m, [
            "Name", "Email", "Role", "Done", "In Progress", "In Review",
            "Story Points", "Docs", "Comments",
        ])
        for row_idx, member in enumerate(members_data.get("members", []), 2):
            ws_m.cell(row=row_idx, column=1, value=member.get("display_name", ""))
            ws_m.cell(row=row_idx, column=2, value=member.get("email", ""))
            ws_m.cell(row=row_idx, column=3, value=member.get("role", ""))
            ws_m.cell(row=row_idx, column=4, value=member.get("done_count", 0))
            ws_m.cell(row=row_idx, column=5, value=member.get("in_progress_count", 0))
            ws_m.cell(row=row_idx, column=6, value=member.get("in_review_count", 0))
            ws_m.cell(row=row_idx, column=7, value=member.get("story_points_sum", 0))
            ws_m.cell(row=row_idx, column=8, value=member.get("docs_count", 0))
            ws_m.cell(row=row_idx, column=9, value=member.get("comments_count", 0))

    if tab in ("projects", "all"):
        ws_p = wb.create_sheet("Projects") if sheets_created else wb.active
        if ws_p is not None and not sheets_created:
            ws_p.title = "Projects"
        elif ws_p is None:
            ws_p = wb.create_sheet("Projects")
        sheets_created = True

        _write_headers(ws_p, [
            "Project Name", "Workspace", "Due Date", "Members",
            "% Completion", "Total", "Done", "In Progress", "In Review",
            "Issue", "Todo", "Unassigned", "Archived", "Status",
        ])
        for row_idx, proj in enumerate(projects_data.get("projects", []), 2):
            ws_p.cell(row=row_idx, column=1, value=proj.get("project_name", ""))
            ws_p.cell(row=row_idx, column=2, value=proj.get("application_name", ""))
            ws_p.cell(row=row_idx, column=3, value=proj.get("due_date", ""))
            ws_p.cell(row=row_idx, column=4, value=", ".join(proj.get("members", [])))
            ws_p.cell(row=row_idx, column=5, value=proj.get("progress_pct", 0))
            ws_p.cell(row=row_idx, column=6, value=proj.get("total", 0))
            ws_p.cell(row=row_idx, column=7, value=proj.get("done", 0))
            ws_p.cell(row=row_idx, column=8, value=proj.get("in_progress", 0))
            ws_p.cell(row=row_idx, column=9, value=proj.get("in_review", 0))
            ws_p.cell(row=row_idx, column=10, value=proj.get("issue", 0))
            ws_p.cell(row=row_idx, column=11, value=proj.get("todo", 0))
            ws_p.cell(row=row_idx, column=12, value=proj.get("unassigned", 0))
            ws_p.cell(row=row_idx, column=13, value=proj.get("archived", 0))
            ws_p.cell(row=row_idx, column=14, value="Archived" if proj.get("is_archived") else "Active")

    if not sheets_created:
        ws_empty = wb.active
        if ws_empty is not None:
            ws_empty.title = "Empty"

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="team_activity_")
    os.close(fd)
    wb.save(path)
    return path


async def generate_export(
    db: AsyncSession,
    app_ids: list[UUID],
    project_ids: list[UUID],
    user_id: UUID,
    date_from: date,
    date_to: date,
    tab: str,
) -> Path:
    """Generate Excel export in a thread pool.

    Args:
        db: Database session.
        app_ids: Owned application IDs.
        project_ids: Scoped project IDs.
        user_id: Requesting user's ID.
        date_from: Start date for the report.
        date_to: End date for the report.
        tab: Which tab to export: "overview", "members", "projects", or "all".

    Returns:
        Path to the generated .xlsx file.
    """
    from fastapi.concurrency import run_in_threadpool

    # Only fetch data needed for the requested tab
    overview = (
        await get_overview(db, project_ids, date_from, date_to)
        if tab in ("overview", "all")
        else None
    )
    members = (
        await get_members_summary(db, app_ids, project_ids, date_from, date_to)
        if tab in ("members", "all")
        else None
    )
    # Only fetch projects data when needed by the requested tab
    projects = (
        await get_projects_summary(db, app_ids, date_from, date_to)
        if tab in ("projects", "all")
        else None
    )

    # Build workbook in thread pool to avoid blocking the event loop
    path_str = await run_in_threadpool(
        _build_workbook,
        overview.model_dump() if overview else {},
        members.model_dump() if members else {},
        projects.model_dump() if projects else {},
        tab,
        date_from.isoformat(),
        date_to.isoformat(),
    )

    return Path(path_str)
